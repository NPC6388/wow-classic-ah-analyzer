import re
import csv
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Border, Side

def convert_price_to_gold(price_str):
    """Convert price from copper to gold.silver.copper format"""
    try:
        copper_total = int(price_str)
        
        # Convert copper to gold, silver, copper
        # 1 gold = 100 silver = 10000 copper
        gold = copper_total // 10000
        remaining_copper = copper_total % 10000
        silver = remaining_copper // 100
        copper = remaining_copper % 100
        
        # Return as decimal gold (gold + silver/100 + copper/10000)
        return gold + (silver / 100) + (copper / 10000)
    except:
        return 0

def format_price_wow(price_gold):
    """Format price for display as WoW format: XgYsZc"""
    try:
        # Convert decimal gold back to copper for formatting
        copper_total = int(price_gold * 10000)
        gold = copper_total // 10000
        remaining_copper = copper_total % 10000
        silver = remaining_copper // 100
        copper = remaining_copper % 100
        return f"{gold}g {silver}s {copper}c"
    except:
        return "0g 0s 0c"

def format_price_copper_to_wow(copper_total):
    """Format copper directly to WoW format: XgYsZc"""
    try:
        copper_total = int(copper_total)
        gold = copper_total // 10000
        remaining_copper = copper_total % 10000
        silver = remaining_copper // 100
        copper = remaining_copper % 100
        return f"{gold}g {silver}s {copper}c"
    except:
        return "0g 0s 0c"

def parse_auc_stat_stddev(file_path):
    """Parse Auc-Stat-StdDev.lua file to get market price data"""
    print(f"Processing StdDev file: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading StdDev file: {e}")
        return {}
    
    market_prices = {}
    
    # Look for item data in format: [12640] = "0:price1;price2;price3;..."
    pattern = r'\[(\d+)\]\s*=\s*"[^:]*:([^"]+)"'
    
    matches = re.finditer(pattern, content)
    
    for match in matches:
        item_id = int(match.group(1))
        price_data = match.group(2)
        
        # Parse the price values (semicolon separated)
        price_parts = price_data.split(';')
        if len(price_parts) >= 1:
            try:
                # Use the most recent price (last entry) as market price
                recent_price_copper = int(price_parts[-1])
                market_price_gold = convert_price_to_gold(recent_price_copper)
                market_prices[item_id] = market_price_gold
                    
            except (ValueError, IndexError):
                continue
    
    print(f"Found market price data for {len(market_prices)} items")
    return market_prices

def parse_auc_stat_simple_market_prices(file_path):
    """Parse Auc-Stat-Simple.lua file to get market price data"""
    print(f"Processing Simple stat file for market prices: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading Simple stat file: {e}")
        return {}
    
    market_prices = {}
    
    # Look for item data in format: ["12640"] = "0@count1;count2;price1;price2;price3;price4"
    # Use the first price field (price1) which matches in-game market prices better
    pattern = r'\["(\d+)"\]\s*=\s*"[^@]*@[^;]*;[^;]*;(\d+\.?\d*);[^"]*"'
    
    matches = re.finditer(pattern, content)
    
    for match in matches:
        item_id = int(match.group(1))
        market_price_copper = float(match.group(2))  # First price value (price1), can be decimal
        
        # Convert market price from copper to gold
        market_price_gold = convert_price_to_gold(market_price_copper)
        market_prices[item_id] = market_price_gold
    
    print(f"Found Simple stat market price data for {len(market_prices)} items")
    return market_prices

def parse_auc_stat_histogram(file_path):
    """Parse Auc-Stat-Histogram.lua file to get times seen and market price data"""
    print(f"Processing histogram file: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading histogram file: {e}")
        return {}, {}
    
    times_seen = {}
    market_prices = {}
    
    # Look for item data in format: ["12640"] = "0@percentile!percentile!price!count!bins;histogram_data"
    pattern = r'\["(\d+)"\]\s*=\s*"[^!]+!([^!]+)!(\d+)!(\d+)!'
    
    matches = re.finditer(pattern, content)
    
    for match in matches:
        item_id = int(match.group(1))
        market_price_copper = int(match.group(3))
        count = int(match.group(4))
        
        # Convert market price from copper to gold
        market_price_gold = convert_price_to_gold(market_price_copper)
        
        # Store the highest count found for each item (in case there are multiple entries)
        if item_id not in times_seen or times_seen[item_id] < count:
            times_seen[item_id] = count
            market_prices[item_id] = market_price_gold
    
    print(f"Found histogram data for {len(times_seen)} items")
    return times_seen, market_prices

def parse_auc_stat_simple(file_path):
    """Parse Auc-Stat-Simple.lua file to get times seen data"""
    print(f"Processing stat file: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading stat file: {e}")
        return {}
    
    times_seen = {}
    
    # Look for item data in format: ["12640"] = "0@something;count;prices..."
    pattern = r'\["(\d+)"\]\s*=\s*"[^;]+;(\d+);'
    
    matches = re.finditer(pattern, content)
    
    for match in matches:
        item_id = int(match.group(1))
        count = int(match.group(2))
        
        # Store the highest count found for each item (in case there are multiple entries)
        if item_id not in times_seen or times_seen[item_id] < count:
            times_seen[item_id] = count
    
    print(f"Found times seen data for {len(times_seen)} items")
    return times_seen

def parse_auctioneer_data(file_path):
    """Parse Auctioneer scan data from Lua file"""
    print(f"Processing: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        print(f"File size: {len(content)} characters")
    except Exception as e:
        print(f"Error reading file: {e}")
        return []
    
    items = []
    
    # Look for the ropes section with the return format
    ropes_match = re.search(r'\["ropes"\]\s*=\s*\{([^}]+(?:\{[^}]*\}[^}]*)*)', content, re.DOTALL)
    
    if ropes_match:
        ropes_content = ropes_match.group(1)
        print(f"Found ropes section, length: {len(ropes_content)}")
        
        # Look for the return statement with the actual data
        return_match = re.search(r'return\s*\{(.+)', ropes_content, re.DOTALL)
        
        if return_match:
            data_content = return_match.group(1)
            print(f"Found return data, length: {len(data_content)}")
            
            # Parse individual item entries using fallback approach immediately
            # Look for item patterns in the data content
            item_pattern = r'Hitem:(\\d+):[^|]*\\|h\\[(.+?)\\]'
            
            matches = re.finditer(item_pattern, data_content, re.DOTALL)
            
            for match in matches:
                try:
                    item_id = int(match.group(1))
                    item_name = match.group(2)
                    
                    # Find the surrounding context for this item
                    start_pos = max(0, match.start() - 50)
                    end_pos = min(len(data_content), match.end() + 200)
                    context = data_content[start_pos:end_pos]
                    
                    # Extract numerical values from the context
                    # Look for pattern: level,quality,count,nil,buyout,bid,time
                    value_pattern = r',(\\d+),(\\d+),(\\d+),nil,(\\d+),(\\d+),(\\d+)'
                    value_match = re.search(value_pattern, context)
                    
                    if not value_match:
                        continue
                        
                    level = int(value_match.group(1))
                    quality = int(value_match.group(2))
                    count = int(value_match.group(3))
                    buyout_price = int(value_match.group(4))
                    bid_price = int(value_match.group(5))
                    time_left = int(value_match.group(6))
                    
                    # Try to find seller name
                    seller_match = re.search(r'\"([^\"]+)\"[^}]*$', context)
                    seller_name = seller_match.group(1) if seller_match and seller_match.group(1) != item_name else "Unknown"
                    
                    # Convert price to gold using our pricing logic
                    buyout_gold = convert_price_to_gold(buyout_price)
                    bid_gold = convert_price_to_gold(bid_price)
                    
                    items.append({
                        'item_id': item_id,
                        'item_name': item_name,
                        'level': level,
                        'quality': quality,
                        'count': count,
                        'buyout_price_copper': buyout_price,
                        'buyout_price_gold': buyout_gold,
                        'bid_price_copper': bid_price,
                        'bid_price_gold': bid_gold,
                        'time_left': time_left,
                        'seller_name': seller_name,
                        'scan_frequency': 1
                    })
                    
                except Exception as e:
                    print(f"Error parsing item: {e} - Match: {match.group(0)[:100] if match else 'None'}")
                    continue
            
            print(f"Found {len(items)} items using precise parsing")
    
    # If precise parsing found few items, try fallback approach
    if len(items) < 10:
        print("Trying fallback parsing approach...")
        
        # Look for any item-like patterns in the entire file
        alt_pattern = r'\|Hitem:(\d+):[^|]*\|h\[([^\]]+)\]\|h\|r'
        
        matches = re.finditer(alt_pattern, content, re.DOTALL)
        
        for match in matches:
            try:
                item_id = int(match.group(1))
                item_name = match.group(2)
                
                # Look for the data after this item link
                item_start = match.start()
                item_end = match.end()
                
                # Get the next 200 characters after the item link
                next_data = content[item_end:item_end+200]
                
                # Look for numeric values
                numbers = re.findall(r'(\d+)', next_data)
                
                if len(numbers) >= 12:
                    level = int(numbers[0])
                    quality = int(numbers[1])
                    count = int(numbers[2])
                    bid_price = int(numbers[3])     # Bid price (lower)
                    buyout_price = int(numbers[11]) # Buyout price (higher)
                    time_left = int(numbers[6])
                    # Position 14 in data structure contains scan frequency
                    scan_frequency = int(numbers[9]) if len(numbers) > 9 else 1
                    
                    
                    # Convert price to gold using our pricing logic
                    buyout_gold = convert_price_to_gold(buyout_price)
                    bid_gold = convert_price_to_gold(bid_price)
                    
                    items.append({
                        'item_id': item_id,
                        'item_name': item_name,
                        'level': level,
                        'quality': quality,
                        'count': count,
                        'buyout_price_copper': buyout_price,
                        'buyout_price_gold': buyout_gold,
                        'bid_price_copper': bid_price,
                        'bid_price_gold': bid_gold,
                        'time_left': time_left,
                        'seller_name': "Unknown",
                        'scan_frequency': scan_frequency
                    })
                    
            except Exception as e:
                continue
        
        print(f"Alternative parsing found {len(items)} items")
    
    print(f"Total items found: {len(items)}")
    return items

def analyze_arbitrage(horde_items, alliance_items, horde_times_seen=None, alliance_times_seen=None, horde_market_prices=None, alliance_market_prices=None):
    """Analyze cross-faction arbitrage opportunities"""
    print("Analyzing arbitrage opportunities...")
    
    # Group items by name
    horde_by_name = defaultdict(list)
    alliance_by_name = defaultdict(list)
    
    for item in horde_items:
        horde_by_name[item['item_name']].append(item)
    
    for item in alliance_items:
        alliance_by_name[item['item_name']].append(item)
    
    # Find items that exist on both factions
    common_items = set(horde_by_name.keys()) & set(alliance_by_name.keys())
    print(f"Found {len(common_items)} items on both factions")
    
    
    arbitrage_opportunities = []
    
    for item_name in common_items:
        horde_items_list = horde_by_name[item_name]
        alliance_items_list = alliance_by_name[item_name]
        
        # Filter out bid-only auctions (no buyout price) first
        horde_buyout_items = [item for item in horde_items_list if item['buyout_price_gold'] > 0]
        alliance_buyout_items = [item for item in alliance_items_list if item['buyout_price_gold'] > 0]
        
        # Skip items that have no buyout auctions on either faction
        if not horde_buyout_items or not alliance_buyout_items:
            continue
            
        # Calculate average prices (for backup if market price not available) - using only buyout auctions
        horde_avg_price = sum(item['buyout_price_gold'] for item in horde_buyout_items) / len(horde_buyout_items)
        alliance_avg_price = sum(item['buyout_price_gold'] for item in alliance_buyout_items) / len(alliance_buyout_items)
        
        # Get market prices from histogram data (prefer this over average)
        item_id = horde_items_list[0]['item_id'] if horde_items_list else alliance_items_list[0]['item_id']
        horde_market_price = horde_market_prices.get(item_id, horde_avg_price) if horde_market_prices else horde_avg_price
        alliance_market_price = alliance_market_prices.get(item_id, alliance_avg_price) if alliance_market_prices else alliance_avg_price
            
        # Find minimum buyout prices (cheapest buyout-only listing on each faction)
        horde_min_price = min(item['buyout_price_gold'] for item in horde_buyout_items)
        alliance_min_price = min(item['buyout_price_gold'] for item in alliance_buyout_items)
        
        
        
        
        # Find price difference between market prices (which will be displayed as buyout prices)
        price_diff = abs(horde_market_price - alliance_market_price)
        
        # Determine which faction is cheaper based on market prices (historic)
        if horde_market_price < alliance_market_price:
            cheaper_historic = "Horde"
        else:
            cheaper_historic = "Alliance"
            
        # Determine which faction has cheaper current buyout price
        if horde_min_price < alliance_min_price:
            cheaper_buyout = "Horde"
        else:
            cheaper_buyout = "Alliance"
        
        # Get times seen from histogram files
        horde_total_scans = horde_times_seen.get(item_id, 0) if horde_times_seen else 0
        alliance_total_scans = alliance_times_seen.get(item_id, 0) if alliance_times_seen else 0
        
        arbitrage_opportunities.append({
            'item_name': item_name,
            'horde_market_price': horde_market_price,
            'alliance_market_price': alliance_market_price,
            'horde_buyout_price': horde_min_price,
            'alliance_buyout_price': alliance_min_price,
            'price_difference': price_diff,
            'cheaper_buyout': cheaper_buyout,
            'cheaper_historic': cheaper_historic,
            'horde_scan_count': horde_total_scans,
            'alliance_scan_count': alliance_total_scans
        })
    
    # Sort by price difference (largest differences first)
    arbitrage_opportunities.sort(key=lambda x: x['price_difference'], reverse=True)
    
    return arbitrage_opportunities

def generate_excel_report(horde_items, alliance_items, arbitrage_opportunities):
    """Generate Excel reports with formatted tables"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # Main analysis sheet
    ws_main = wb.active
    ws_main.title = "Arbitrage Analysis"
    
    # Headers for main analysis
    headers = [
        'Item Name', 'Times Seen (Horde, Alliance)', 'Horde Buyout Price', 
        'Alliance Buyout Price', 'Price Difference', 'Horde Market Price', 
        'Alliance Market Price', 'Cheaper Buyout', 'Cheaper Historic'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws_main.cell(row=1, column=col, value=header)
    
    # Write arbitrage opportunities
    for row, opp in enumerate(arbitrage_opportunities, 2):
        price_diff_gold = int(round(opp['price_difference']))
        
        ws_main.cell(row=row, column=1, value=opp['item_name'])
        ws_main.cell(row=row, column=2, value=f"{opp['horde_scan_count']}, {opp['alliance_scan_count']}")
        ws_main.cell(row=row, column=3, value=format_price_wow(opp['horde_buyout_price']))
        ws_main.cell(row=row, column=4, value=format_price_wow(opp['alliance_buyout_price']))
        ws_main.cell(row=row, column=5, value=f"{price_diff_gold}g")
        ws_main.cell(row=row, column=6, value=format_price_wow(opp['horde_market_price']))
        ws_main.cell(row=row, column=7, value=format_price_wow(opp['alliance_market_price']))
        ws_main.cell(row=row, column=8, value=opp['cheaper_buyout'])
        ws_main.cell(row=row, column=9, value=opp['cheaper_historic'])
    
    # Convert main sheet to table
    if len(arbitrage_opportunities) > 0:
        table_ref = f"A1:I{len(arbitrage_opportunities) + 1}"
        table = Table(displayName="ArbitrageTable", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws_main.add_table(table)
    
    # Add conversion calculator below the table
    calculator_start_row = len(arbitrage_opportunities) + 4  # Leave some space after table
    
    # Calculator title
    ws_main.cell(row=calculator_start_row, column=1, value="Per-Unit Price Calculator")
    ws_main.cell(row=calculator_start_row, column=1).font = Font(bold=True, size=14)
    
    # Calculator labels and input fields
    ws_main.cell(row=calculator_start_row + 2, column=1, value="Stack Size:")
    ws_main.cell(row=calculator_start_row + 2, column=2).border = Border(outline=Side(style='thin'))
    ws_main.cell(row=calculator_start_row + 2, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    ws_main.cell(row=calculator_start_row + 3, column=1, value="Total Buyout Price (copper):")
    ws_main.cell(row=calculator_start_row + 3, column=2).border = Border(outline=Side(style='thin'))
    ws_main.cell(row=calculator_start_row + 3, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Result field with formula
    ws_main.cell(row=calculator_start_row + 5, column=1, value="Per-Unit Price (copper):")
    ws_main.cell(row=calculator_start_row + 5, column=1).font = Font(bold=True)
    
    # Formula to calculate per-unit price (Total Price / Stack Size)
    stack_cell = f"B{calculator_start_row + 2}"
    price_cell = f"B{calculator_start_row + 3}"
    formula_cell = f"B{calculator_start_row + 5}"
    ws_main.cell(row=calculator_start_row + 5, column=2, value=f"=IF(AND({stack_cell}<>0,{stack_cell}<>\"\",{price_cell}<>\"\"),{price_cell}/{stack_cell},\"Enter values above\")")
    ws_main.cell(row=calculator_start_row + 5, column=2).font = Font(bold=True)
    ws_main.cell(row=calculator_start_row + 5, column=2).border = Border(outline=Side(style='thick'))
    ws_main.cell(row=calculator_start_row + 5, column=2).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # Gold conversion
    ws_main.cell(row=calculator_start_row + 6, column=1, value="Per-Unit Price (gold):")
    ws_main.cell(row=calculator_start_row + 6, column=1).font = Font(bold=True)
    ws_main.cell(row=calculator_start_row + 6, column=2, value=f"=IF(ISNUMBER({formula_cell}),{formula_cell}/10000,\"\")")
    ws_main.cell(row=calculator_start_row + 6, column=2).font = Font(bold=True)
    ws_main.cell(row=calculator_start_row + 6, column=2).border = Border(outline=Side(style='thick'))
    ws_main.cell(row=calculator_start_row + 6, column=2).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws_main.cell(row=calculator_start_row + 6, column=2).number_format = '0.0000'
    
    # Instructions
    ws_main.cell(row=calculator_start_row + 8, column=1, value="Instructions: Enter the stack size and total buyout price to calculate per-unit pricing")
    ws_main.cell(row=calculator_start_row + 8, column=1).font = Font(italic=True, color="666666")
    
    # Auto-adjust column widths
    for col in ws_main.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_main.column_dimensions[column].width = adjusted_width
    
    # Horde bargains sheet
    ws_horde = wb.create_sheet("Horde Bargains")
    horde_headers = ['Item Name', 'Buyout Price (Gold)', 'Count', 'Seller']
    
    for col, header in enumerate(horde_headers, 1):
        ws_horde.cell(row=1, column=col, value=header)
    
    # Group Horde items by name and find lowest prices
    horde_by_name = defaultdict(list)
    for item in horde_items:
        horde_by_name[item['item_name']].append(item)
    
    horde_bargains = []
    for name, items in horde_by_name.items():
        # Filter to only buyout auctions
        buyout_items = [item for item in items if item['buyout_price_gold'] > 0]
        if buyout_items:  # Only add if there are buyout auctions
            min_price_item = min(buyout_items, key=lambda x: x['buyout_price_gold'])
            horde_bargains.append({
                'name': name,
                'price': min_price_item['buyout_price_gold'],
                'count': min_price_item['count'],
                'seller': min_price_item['seller_name']
            })
    
    horde_bargains.sort(key=lambda x: x['price'])
    for row, item in enumerate(horde_bargains[:100], 2):
        ws_horde.cell(row=row, column=1, value=item['name'])
        ws_horde.cell(row=row, column=2, value=format_price_wow(item['price']))
        ws_horde.cell(row=row, column=3, value=item['count'])
        ws_horde.cell(row=row, column=4, value=item['seller'])
    
    # Convert Horde sheet to table
    if len(horde_bargains) > 0:
        horde_table_ref = f"A1:D{min(101, len(horde_bargains) + 1)}"
        horde_table = Table(displayName="HordeBargainsTable", ref=horde_table_ref)
        horde_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        horde_table.tableStyleInfo = horde_style
        ws_horde.add_table(horde_table)
    
    # Auto-adjust Horde column widths
    for col in ws_horde.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_horde.column_dimensions[column].width = adjusted_width
    
    # Alliance bargains sheet
    ws_alliance = wb.create_sheet("Alliance Bargains")
    alliance_headers = ['Item Name', 'Buyout Price (Gold)', 'Count', 'Seller']
    
    for col, header in enumerate(alliance_headers, 1):
        ws_alliance.cell(row=1, column=col, value=header)
    
    # Group Alliance items by name and find lowest prices
    alliance_by_name = defaultdict(list)
    for item in alliance_items:
        alliance_by_name[item['item_name']].append(item)
    
    alliance_bargains = []
    for name, items in alliance_by_name.items():
        # Filter to only buyout auctions
        buyout_items = [item for item in items if item['buyout_price_gold'] > 0]
        if buyout_items:  # Only add if there are buyout auctions
            min_price_item = min(buyout_items, key=lambda x: x['buyout_price_gold'])
            alliance_bargains.append({
                'name': name,
                'price': min_price_item['buyout_price_gold'],
                'count': min_price_item['count'],
                'seller': min_price_item['seller_name']
            })
    
    alliance_bargains.sort(key=lambda x: x['price'])
    for row, item in enumerate(alliance_bargains[:100], 2):
        ws_alliance.cell(row=row, column=1, value=item['name'])
        ws_alliance.cell(row=row, column=2, value=format_price_wow(item['price']))
        ws_alliance.cell(row=row, column=3, value=item['count'])
        ws_alliance.cell(row=row, column=4, value=item['seller'])
    
    # Convert Alliance sheet to table
    if len(alliance_bargains) > 0:
        alliance_table_ref = f"A1:D{min(101, len(alliance_bargains) + 1)}"
        alliance_table = Table(displayName="AllianceBargainsTable", ref=alliance_table_ref)
        alliance_style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                                      showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        alliance_table.tableStyleInfo = alliance_style
        ws_alliance.add_table(alliance_table)
    
    # Auto-adjust Alliance column widths
    for col in ws_alliance.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_alliance.column_dimensions[column].width = adjusted_width
    
    # Save Excel file
    excel_filename = f"ah_analysis_{timestamp}.xlsx"
    wb.save(excel_filename)
    
    print(f"Generated Excel report: {excel_filename}")
    print(f"- Arbitrage Analysis sheet with {len(arbitrage_opportunities)} opportunities")
    print(f"- Horde Bargains sheet with top 100 items")
    print(f"- Alliance Bargains sheet with top 100 items")
    
    return excel_filename

def main():
    """Main function"""
    print("WoW Classic AH Analyzer - Final Version")
    print("=" * 50)
    
    # File paths
    horde_path = r"C:\Program Files (x86)\World of Warcraft\_classic_era_\WTF\Account\LUCIA1\SavedVariables\auc-scandata.lua"
    alliance_path = r"C:\Program Files (x86)\World of Warcraft\_classic_era_\WTF\Account\51718250#1\SavedVariables\auc-scandata.lua"
    
    # Histogram file paths (contains times seen)
    horde_histogram_path = r"C:\Program Files (x86)\World of Warcraft\_classic_era_\WTF\Account\LUCIA1\SavedVariables\Auc-Stat-Histogram.lua"
    alliance_histogram_path = r"C:\Program Files (x86)\World of Warcraft\_classic_era_\WTF\Account\51718250#1\SavedVariables\Auc-Stat-Histogram.lua"
    
    # Simple stat file paths (contains market prices)
    horde_simple_path = r"C:\Program Files (x86)\World of Warcraft\_classic_era_\WTF\Account\LUCIA1\SavedVariables\Auc-Stat-Simple.lua"
    alliance_simple_path = r"C:\Program Files (x86)\World of Warcraft\_classic_era_\WTF\Account\51718250#1\SavedVariables\Auc-Stat-Simple.lua"
    
    print(f"Horde data path: {horde_path}")
    print(f"Alliance data path: {alliance_path}")
    print(f"Horde histogram path: {horde_histogram_path}")
    print(f"Alliance histogram path: {alliance_histogram_path}")
    print(f"Horde Simple stat path: {horde_simple_path}")
    print(f"Alliance Simple stat path: {alliance_simple_path}")
    print()
    
    # Check if files exist
    if not os.path.exists(horde_path):
        print(f"ERROR: Horde scan data not found at {horde_path}")
        return
    
    if not os.path.exists(alliance_path):
        print(f"ERROR: Alliance scan data not found at {alliance_path}")
        return
    
    # Extract times seen data and market prices from histogram files
    print("Extracting Horde histogram data...")
    horde_times_seen, horde_histogram_market_prices = parse_auc_stat_histogram(horde_histogram_path)
    
    print("Extracting Alliance histogram data...")
    alliance_times_seen, alliance_histogram_market_prices = parse_auc_stat_histogram(alliance_histogram_path)
    
    # Extract market price data from Simple stat files (better than histogram prices)
    print("Extracting Horde market price data...")
    horde_market_prices = parse_auc_stat_simple_market_prices(horde_simple_path)
    
    print("Extracting Alliance market price data...")
    alliance_market_prices = parse_auc_stat_simple_market_prices(alliance_simple_path)
    
    # Also get times seen from Simple stat (more accurate than histogram)
    print("Extracting Horde times seen from Simple stat...")
    horde_simple_times_seen = parse_auc_stat_simple(horde_simple_path)
    
    print("Extracting Alliance times seen from Simple stat...")
    alliance_simple_times_seen = parse_auc_stat_simple(alliance_simple_path)
    
    # Use Simple stat times seen (they're more accurate)
    horde_times_seen = horde_simple_times_seen
    alliance_times_seen = alliance_simple_times_seen
    
    # Extract auction data
    print("Extracting Horde auction data...")
    horde_items = parse_auctioneer_data(horde_path)
    
    print("Extracting Alliance auction data...")
    alliance_items = parse_auctioneer_data(alliance_path)
    
    if not horde_items and not alliance_items:
        print("No auction data found in either file!")
        print("This might indicate the data format is different than expected.")
        print("Please check that the files contain Auctioneer scan data.")
        return
    
    print(f"\nTotal items found:")
    print(f"Horde: {len(horde_items)}")
    print(f"Alliance: {len(alliance_items)}")
    
    if len(horde_items) == 0 and len(alliance_items) == 0:
        print("No items found. The data format may be different than expected.")
        return
    
    # Analyze arbitrage opportunities
    arbitrage_opportunities = analyze_arbitrage(horde_items, alliance_items, horde_times_seen, alliance_times_seen, horde_market_prices, alliance_market_prices)
    
    if len(arbitrage_opportunities) == 0:
        print("No arbitrage opportunities found.")
        print("This could mean:")
        print("- No items exist on both factions")
        print("- Price differences are too small")
        print("- Data parsing needs adjustment")
    
    # Generate reports
    excel_file = generate_excel_report(horde_items, alliance_items, arbitrage_opportunities)
    
    print(f"\nAnalysis complete!")
    print(f"Excel report: {excel_file}")
    
    # Open the Excel file
    try:
        os.startfile(excel_file)
        print("Opened Excel report")
    except:
        print(f"Please open {excel_file} manually")

if __name__ == "__main__":
    main()

